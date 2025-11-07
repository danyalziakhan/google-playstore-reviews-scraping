from glob import glob
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

if __name__ == "__main__":
    files = glob("output/**/*.csv", recursive=True)
    df = pd.concat((pd.read_csv(file) for file in files), ignore_index=True)
    df = df.drop_duplicates()
    df = df.drop(columns=["uid"])
    df.to_excel("REVIEWS_SCRAPED.xlsx", index=False)

    # Load workbook and active sheet
    wb = load_workbook("REVIEWS_SCRAPED.xlsx")
    ws = wb.active

    # Adjust width only based on header text
    for i, col_name in enumerate(df.columns, start=1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = (  # type: ignore
            len(str(col_name)) + 2
        )  # add padding

    wb.save("REVIEWS_SCRAPED.xlsx")
