import asyncio
from glob import glob
import os
import re
from playwright.async_api import async_playwright, Page
import pandas as pd
import dataclasses
import argparse
from log import info, warning
from playwright_stealth import Stealth
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


@dataclasses.dataclass
class Review:
    product_name: str = ""
    username: str = ""
    review_date: str = ""
    rating: str = ""
    review_content: str = ""
    uid: str = ""


WAIT_FOR_TIMEOUT = 100000
RELOAD_TIMEOUT = 10000

rename_map = {
    "product_name": "Product Name",
    "username": "Username",
    "review_date": "Review Date",
    "rating": "Rating",
    "review_content": "Review Content",
    "uid": "uid",
}


async def run(
    headless: bool,
    skip_rows: list[str],
):
    async with Stealth().use_async(async_playwright()) as p:
        chrome_path = os.path.join(
            os.environ["LOCALAPPDATA"], "Google", "Chrome", "User Data", "Default"
        )
        browser = await p.chromium.launch_persistent_context(
            user_data_dir=chrome_path,
            headless=headless,
            executable_path=r"C:\Program Files\Google\Chrome\Application\chrome.exe",
            args=["--disable-blink-features=AutomationControlled"],
        )

        page = await browser.new_page()

        await extract_reviews(page, [int(s) for s in skip_rows])

        await browser.close()

        files = glob("output/**/*.csv", recursive=True)
        df = pd.concat((pd.read_csv(file) for file in files), ignore_index=True)
        df = df.drop_duplicates()
        df = df.drop(columns=["uid"])
        df.to_excel("REVIEWS_SCRAPED.xlsx", index=False)

        # Load workbook and active sheet
        wb = load_workbook("REVIEWS_SCRAPED.xlsx")
        ws = wb.active

        # Adjust width only based on header text
        for i, col_name in enumerate(df.columns, start=1):
            col_letter = get_column_letter(i)
            ws.column_dimensions[col_letter].width = (  # type: ignore
                len(str(col_name)) + 2
            )  # add padding

        wb.save("REVIEWS_SCRAPED.xlsx")


def load_seen_uids(idx: int):
    base_dir = "output"
    seen = set()

    if not os.path.exists(base_dir):
        return seen

    csv_files = glob(os.path.join(base_dir, f"{idx}_batch*.csv"))
    for file_path in csv_files:
        try:
            df = pd.read_csv(file_path, usecols=["uid"])
            seen.update(df["uid"].dropna().astype(str))
        except Exception as e:
            print(f"Skipping {file_path}: {e}")

    print(f"Loaded {len(seen)} previously seen UIDs from saved files.")
    return seen


def extract_rating(text):
    match = re.search(r"(\d+(?:\.\d+)?)\s*stars?", text)
    return float(match.group(1)) if match else None


async def extract_reviews(
    page: Page,
    skip_rows: list[int],
):
    df = pd.read_excel("Products.xlsx")
    product_urls = df["Product URL"].tolist()

    for idx, product_url in enumerate(product_urls):
        if (idx + 1) in skip_rows:
            continue

        info(f"Product URL: {product_url}")

        await visit_product_link(page, product_url, wait_until="load")

        product_name = await page.text_content('h1 > span[itemprop="name"]')

        if not product_name:
            raise ValueError("Product Name Text not found")

        more_reviews_button = await page.query_selector(
            "span:has-text('See all reviews')"
        )
        if not more_reviews_button:
            raise ValueError("More Reviews Button not found")

        await more_reviews_button.click()

        await page.wait_for_selector(
            f"div[aria-label='{product_name}'] > div > div > div > div.fysCi.Vk3ZVd > div > div:nth-child(2) > div"
        )

        review_results = []

        seen_from_disk = load_seen_uids(idx)
        review_idx = len(seen_from_disk)

        async for review, uid in iter_reviews(
            page, product_name, external_seen=seen_from_disk
        ):
            review_idx += 1

            info(f"Scraping {review_idx} review")
            review_scraped = Review()
            review_scraped.product_name = product_name
            review_scraped.uid = uid

            username = await review.query_selector(
                "header > div > div:nth-child(1) > div"
            )

            if not username:
                continue

            username = await username.text_content()

            if not username:
                raise ValueError("Username Text not found")

            username = username.strip()

            info(f"Username: {username}")

            review_scraped.username = username

            date = await review.query_selector("header > div.Jx4nYe > span")

            if not date:
                continue

            date = await date.text_content()

            if not date:
                raise ValueError("Date Text not found")

            date = date.strip()
            try:
                normalized = datetime.strptime(date, "%B %d, %Y").strftime("%Y-%m-%d")
                info(f"Date: {normalized}")
                review_scraped.review_date = normalized
            except ValueError:
                info(f"Date: {date}")
                review_scraped.review_date = date

            ratings = await review.query_selector("header > div.Jx4nYe > div")

            if not ratings:
                continue

            ratings = await ratings.get_attribute("aria-label")

            if not ratings:
                raise ValueError("Ratings Text not found")

            ratings = str(extract_rating(ratings.strip()))
            info(f"Ratings: {ratings}")

            review_scraped.rating = ratings

            try:
                review_content = await review.query_selector("div.h3YV2d")
            except Exception:
                continue

            if not review_content:
                continue

            review_content = await review_content.text_content()

            if not review_content:
                raise ValueError("Review context Text not found")

            review_content = review_content.strip()
            review_scraped.review_content = review_content
            review_results.append(review_scraped)

            # Save every 200 reviews
            if review_idx % 200 == 0:
                df = pd.DataFrame(review_results)
                df = df.rename(columns=rename_map)[list(rename_map.values())]
                file_path = os.path.join("output", f"{idx}_batch_{review_idx}.csv")

                # Append if file exists, else create new
                if os.path.exists(file_path):
                    df.to_csv(file_path, mode="a", index=False, header=False)
                else:
                    df.to_csv(file_path, mode="w", index=False, header=True)

                info(f"Saved batch {review_idx} reviews.")
                review_results.clear()

        if review_results:
            df = pd.DataFrame(review_results)
            df = df.rename(columns=rename_map)[list(rename_map.values())]

            os.makedirs(os.path.join("output"), exist_ok=True)
            file_path = os.path.join(
                "output",
                f"{idx}_batch_last.csv",
            )

            # Append if file exists, else create new
            if os.path.exists(file_path):
                df.to_csv(file_path, mode="a", index=False, header=False)
            else:
                df.to_csv(file_path, mode="w", index=False, header=True)


async def visit_product_link(page: Page, product_url: str, wait_until):
    # * Retry visiting the link 10 times
    for retry in range(1, 11):
        try:
            await page.goto(
                product_url,
                wait_until=wait_until,
            )
        except Exception:
            warning(f"Page didn't load at all. Retrying # {retry} ... ")
            await page.wait_for_timeout(1000)
        else:
            break
    else:
        raise ValueError("Page didn't load even after a retry")


async def iter_reviews(
    page: Page,
    product_name: str,
    limit=1000,
    max_no_new=500,
    poll_interval=0.02,
    external_seen=None,
):
    seen = set(external_seen or [])
    url = page.url

    # Dynamically increase allowed scroll attempts when resuming
    if external_seen:
        max_no_new = max(max_no_new, int(len(external_seen) * 0.05))
        info(
            f"Adjusted max_no_new to {max_no_new} based on {len(external_seen)} seen reviews."
        )

    async def reopen_reviews_tab():
        try:
            await page.wait_for_load_state("networkidle")
        except Exception:
            pass
        try:
            await page.click("span:has-text('See all reviews')")
        except Exception:
            pass
        await page.wait_for_timeout(1000)

    info(f"Total reviews to fetch: {limit}")
    await page.wait_for_timeout(1000)

    selector = f"div[aria-label='{product_name}'] > div > div > div > div.fysCi.Vk3ZVd > div > div:nth-child(2) > div"

    total_fetched = len(seen)
    no_new_counter = 0

    while len(seen) < limit:
        await page.evaluate("""
        () => {
        const el = document.querySelector('.fysCi.Vk3ZVd');
        if (el) el.scrollBy(0, 1200);
        }
        """)
        await page.wait_for_timeout(int(poll_interval * 1000))

        try:
            elements = await page.query_selector_all(selector)
        except Exception:
            continue

        new_found = False
        for el in elements:
            try:
                uid = await el.evaluate("""el => el.outerHTML.slice(0,200)""")
            except Exception:
                continue

            if not uid:
                try:
                    uid = await el.evaluate(
                        "el => el.outerHTML.slice(0,200) + '|' + el.getBoundingClientRect().top"
                    )
                except Exception:
                    continue

            if uid not in seen:
                seen.add(uid)
                new_found = True
                total_fetched += 1
                yield el, uid
                if total_fetched >= limit:
                    return

        if new_found:
            no_new_counter = 0
        else:
            no_new_counter += 1

        # Step 2: If page is hanging, refresh
        if no_new_counter >= max_no_new:
            info(f"Reloading page to continue... {total_fetched}/{limit}")
            await page.goto(url)
            await reopen_reviews_tab()

            # Step 3: Scroll down until a new unseen review appears
            info("Scrolling back to where we left off...")
            resume_attempts = 0
            while resume_attempts < 100:  # safety limit
                await page.evaluate("""
                () => {
                const el = document.querySelector('.fysCi.Vk3ZVd');
                if (el) el.scrollBy(0, 1200);
                }
                """)
                await page.wait_for_timeout(int(poll_interval * 1000))

                try:
                    elements = await page.query_selector_all(selector)
                except Exception:
                    continue

                found_unseen = False
                for el in elements:
                    try:
                        uid = await el.evaluate("""el => el.outerHTML.slice(0,200)""")
                    except Exception:
                        continue

                    if uid and uid not in seen:
                        found_unseen = True
                        break

                if found_unseen:
                    info("Resumed at next unseen review.")
                    break

                resume_attempts += 1

            no_new_counter = 0


def comma_separated_list(value):
    return [v.strip() for v in value.split(",") if v.strip()]


if __name__ == "__main__":
    os.makedirs("output", exist_ok=True)

    parser = argparse.ArgumentParser(description="Google Play Store Reviews Scraper")
    parser.add_argument(
        "--headless",
        action="store_true",
        help="Headless mode",
    )
    parser.add_argument(
        "--skip_rows",
        type=comma_separated_list,
        help="Rows to skip",
    )

    args = parser.parse_args()

    asyncio.run(
        run(
            args.headless,
            args.skip_rows if args.skip_rows else [],
        )
    )
